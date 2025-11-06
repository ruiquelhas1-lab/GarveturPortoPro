# tools/xlsx2geojson.py
"""
Conversor XLSX -> GeoJSON/JSON (auto-deteta linha de cabeçalhos; anti-NaN).
- Lê data/empreendimentos.xlsx (folha 1).
- Procura a linha de cabeçalho nas primeiras 10 linhas (nome/latitude/longitude).
- Cabeçalhos aceites (case-insensitive; com/sem acentos): 
  nome, concelho, estado, promotor, tipologias, conclusao, preco_m2, notas, latitude, longitude, link
- Gera data/empreendimentos.json (FeatureCollection) SEM NaN/Inf (substitui por None).
"""
import os, sys, math, json, unicodedata

XLSX_IN = "data/empreendimentos.xlsx"
JSON_OUT = "data/empreendimentos.json"

try:
    import openpyxl
except Exception:
    print("ERRO: openpyxl não está disponível. Instale com 'pip install openpyxl'.", file=sys.stderr)
    sys.exit(1)

REQUIRED = ["nome","latitude","longitude"]
OPTIONAL = ["concelho","estado","promotor","tipologias","conclusao","preco_m2","notas","link"]
ALL = REQUIRED + OPTIONAL

def norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s

def to_num(v):
    if v is None: return None
    s = str(v).strip().replace(",", ".")
    try:
        n = float(s)
        if math.isnan(n) or math.isinf(n):
            return None
        return n
    except:
        return None

def sanitize(o):
    if isinstance(o, float):
        if math.isnan(o) or math.isinf(o):
            return None
        return o
    if isinstance(o, dict):
        return {k: sanitize(v) for k,v in o.items()}
    if isinstance(o, list):
        return [sanitize(v) for v in o]
    return o

if not os.path.exists(XLSX_IN):
    print(f"ERRO: não encontrei o ficheiro {XLSX_IN}", file=sys.stderr)
    sys.exit(1)

wb = openpyxl.load_workbook(XLSX_IN, data_only=True)
ws = wb.worksheets[0]

# 1) Encontrar a linha de cabeçalho nas primeiras 10 linhas
def find_header_row():
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        vals = [norm(c if c is not None else "") for c in row]
        if not any(vals):  # linha vazia
            continue
        # requisitos mínimos: algum "nome" e algum "lat" e algum "lon"
        has_nome = any("nome" == v or "empreendimento" in v or "projeto" in v or "projecto" in v for v in vals)
        has_lat = any(v in ("latitude","lat") for v in vals)
        has_lon = any(v in ("longitude","lng","long","lon") for v in vals)
        if has_nome and has_lat and has_lon:
            return row_idx, [norm(c if c is not None else "") for c in row]
    return None, None

header_row, headers_norm = find_header_row()
if not header_row:
    print("ERRO: não encontrei linha de cabeçalhos nas primeiras 10 linhas (preciso de nome/latitude/longitude).", file=sys.stderr)
    sys.exit(1)

idx = { h:i for i,h in enumerate(headers_norm) }

aliases = {
  "nome": ["nome", "empreendimento", "projeto", "projecto", "name", "titulo", "título"],
  "concelho": ["concelho","municipio","município","city","local"],
  "estado": ["estado","status","fase"],
  "promotor": ["promotor","developer","owner"],
  "tipologias": ["tipologias","tipos","typologies"],
  "conclusao": ["conclusao","conclusão","entrega","completion","ano","ano_conclusao"],
  "preco_m2": ["preco_m2","preco_med_m2","preco_medio_m2","preço_m2","price_m2","price_per_sqm"],
  "notas": ["notas","observacoes","observações","notes"],
  "latitude": ["latitude","lat"],
  "longitude": ["longitude","lng","long","lon","x","y"],
  "link": ["link","url","website"]
}

colmap = {}
for key, cands in aliases.items():
    for cand in cands:
        n = norm(cand)
        if n in idx:
            colmap[key] = idx[n]
            break

missing_required = [k for k in REQUIRED if k not in colmap]
if missing_required:
    print("ERRO: faltam colunas obrigatórias (nome/latitude/longitude).", file=sys.stderr)
    print("Cabeçalhos detetados (linha", header_row, "):", headers_norm, file=sys.stderr)
    sys.exit(1)

# Log útil
print("Linha de cabeçalho detetada:", header_row)
print("Mapeamento de colunas:", {k: list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0][v] for k,v in colmap.items()})

feats, rows, skipped = [], 0, 0
for r in ws.iter_rows(min_row=header_row+1, values_only=True):
    rows += 1
    def get(col):
        j = colmap.get(col)
        return r[j] if j is not None and j < len(r) else None

    lat = to_num(get("latitude"))
    lng = to_num(get("longitude"))
    nome_raw = get("nome")
    nome = (str(nome_raw).strip() if nome_raw is not None else "")

    if not nome or lat is None or lng is None:
        skipped += 1
        continue

    props = {}
    for k in ALL:
        if k in ("latitude","longitude","nome"): continue
        v = get(k)
        props[k] = "" if v is None else str(v).strip()

    feats.append({
        "type": "Feature",
        "properties": {"nome": nome, **props},
        "geometry": {"type":"Point", "coordinates":[lng, lat]}
    })

gj = {"type":"FeatureCollection", "features":feats}
gj = sanitize(gj)

os.makedirs(os.path.dirname(JSON_OUT), exist_ok=True)
with open(JSON_OUT, "w", encoding="utf-8") as f:
    json.dump(gj, f, ensure_ascii=False, indent=2, allow_nan=False)

print(f"OK: {len(feats)} pontos gerados. Linhas lidas: {rows} (ignoradas por falta de coords/nome: {skipped}).")
